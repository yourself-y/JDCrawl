# -*- coding:utf-8 -*-
import openpyxl


class Outputer(object):
    def collect_data(self, phone_model, evaluation_number, price, merchant):
        print '^ output'
        pathname = 'C:/Users/YYT/Desktop/JD.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'phoneData'
        ws.merge_cells('A1:E1')
        ws['A1'] = '京东商城手机信息采集'
        ws['A2'] = '序号'
        ws['B2'] = '手机型号'
        ws['C2'] = '评价人数'
        ws['D2'] = '价格'
        ws['E2'] = '商家'

        print len(phone_model)
        print len(evaluation_number)
        print len(merchant)
        for row in range(0, len(phone_model)):
            ws.cell(row=(row + 3), column=1).value = row + 1
            ws.cell(row=(row + 3), column=2).value = phone_model[row]
            ws.cell(row=(row + 3), column=3).value = evaluation_number[row]
            ws.cell(row=(row + 3), column=4).value = price[row]
            ws.cell(row=(row + 3), column=5).value = merchant[row]

        wb.save(filename=pathname)
        wb.close()

